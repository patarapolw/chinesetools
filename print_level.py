from openpyxl import load_workbook
from source.common import *

class Level():
	def __init__(self):
		with open('database/level.txt', 'r', encoding='utf8') as f:
			read = False
			self.hanzi = []
			for line in f.readlines():
				if line[:-1] == 'SpoonFed':
					read = True
					continue
				if line == '\n':
					read = False
				if read:
					self.hanzi += [line[:-1]]

	def getLevelArray(self, vocab):
		self.level_array = []
		for char in vocab:
			set_level = False
			for level in range(len(self.hanzi)):
				if char in self.hanzi[level]:
					self.level_array += [level+1]
					set_level = True
					break

			if not set_level:
				self.level_array += [100]
		return self.level_array

	def getLevel(self, vocab):
		return max(self.getLevelArray(vocab))

filename = 'database/HSK.xlsx'
wb = load_workbook(filename)

sheet = wb['Vocab']
col = 'J'

sheet[col + str(1)] = 'SpoonFed Level'

level = Level()
printAnything(level.hanzi)

for row in range(2, sheet.max_row+1):
	vocab = sheet['A' + str(row)].value
	sheet[col + str(row)] = level.getLevel(vocab)

	printText(vocab)
	print(level.level_array)

wb.save(filename)