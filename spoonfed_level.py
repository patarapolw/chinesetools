from openpyxl import load_workbook
from source.common import *
import re

filename = 'database/SpoonFed.xlsx'
wb = load_workbook(filename)

sheet = wb['SpoonFed']

hanzi = ''

for row in range(2, sheet.max_row+1):
	sentence = sheet['C' + str(row)].value
	for char in re.findall('[\u2E80-\u2FD5\u3400-\u4DBF\u4E00-\u9FCC]',sentence):
		if char in hanzi:
			continue
		hanzi += char

batch = len(hanzi)//60 + 1
level = []
for start in range(0, len(hanzi), batch):
	level += [hanzi[start:start+batch]]

for string in level:
	printText(string)
